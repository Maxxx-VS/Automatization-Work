import openpyxl
from openpyxl.styles import PatternFill

general_arry = []
greenFill = PatternFill(start_color='99EE99', end_color='99EE99', fill_type='solid')

def first():
    # открываем excel-файл и выбираем 1 страницу
    wb1 = openpyxl.load_workbook(filename = '1.xlsx')
    sheet = wb1['Лист1']
    count_rows = sheet['A']
 
    for i in range(3, len(count_rows)+1):
        row_A = sheet[f'A{i-1}'].value
        row_T = sheet[f'T{i}'].value

        # проверяем полное прохождение по маршруту и добавляем готовые марки в список
        try:
            if row_T is not None and row_T != 'Дата':
                row_spl = row_A.split()
                if len(row_spl) == 3:
                    general_arry.append(row_spl[-1])
        except AttributeError:
            continue

    # print(general_arry)
    # print(len(general_arry))

def second():
    # открываем excel-файл и выбираем 2 страницу
    wb2 = openpyxl.load_workbook(filename = '2.xlsx')
    sheet = wb2['Лист1']
    count_rows_2 = sheet['A']

    for i in range(3, len(count_rows_2)):
        row_B = sheet[f'B{i}'].value
        row_B_color = sheet[f'B{i}']

        try:
            for j in general_arry:
                if row_B == j:
                    row_B_color.fill = greenFill

        except AttributeError:
            continue

    wb2.save('3.xlsx')


first()
second()
print("ГОТОВО!")

    
